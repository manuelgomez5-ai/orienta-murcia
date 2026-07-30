from __future__ import annotations
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "input" / "ranking.xlsx"
OUTPUT = ROOT / "data" / "ranking.json"


def main() -> None:
    if not INPUT.exists():
        print("No hay input/ranking.xlsx; se mantiene el ranking existente.")
        return
    wb = load_workbook(INPUT, data_only=True, read_only=True)
    ws = wb["Ranking completo"] if "Ranking completo" in wb.sheetnames else wb.active
    header_row = None
    headers = {}
    for row in ws.iter_rows():
        values = [str(c.value).strip() if c.value is not None else "" for c in row]
        if "Puesto global" in values and "Nº lista" in values:
            header_row = row[0].row
            headers = {v: i for i, v in enumerate(values) if v}
            break
    if header_row is None:
        raise RuntimeError("No se encontró la cabecera del ranking.")
    rows = []
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        try:
            global_pos = int(values[headers["Puesto global"]])
            block = int(values[headers["Bloque"]])
            block_rank = int(values[headers["Puesto bloque"]])
        except (TypeError, ValueError):
            continue
        rows.append({
            "global": global_pos,
            "block": block,
            "block_rank": block_rank,
            "list_number": str(values[headers["Nº lista"]] or "").strip(),
            "masked_id": str(values[headers["DNI"]] or "").strip(),
            "name": str(values[headers.get("Apellidos, nombre", headers.get("Apellidos y nombre"))] or "").strip(),
            "points": float(values[headers["Puntos"]] or 0),
        })
    block1 = sum(1 for r in rows if r["block"] == 1)
    block2 = sum(1 for r in rows if r["block"] == 2)
    payload = {
        "status": "complete",
        "course": "2026/2027",
        "source_label": INPUT.name,
        "official_total": len(rows),
        "block_1_total": block1,
        "block_2_total": block2,
        "notice": "Ranking completo importado.",
        "rows": rows,
    }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Importadas {len(rows)} personas.")


if __name__ == "__main__":
    main()
